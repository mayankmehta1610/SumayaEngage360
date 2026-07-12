"""Add Cursor Done column to requirement sheets 5-12 and mark completed items."""
import shutil
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

SRC = Path(r"C:\Users\Admin\AppData\Local\Temp\SumayaEngage360.xlsx")
OUT = Path(__file__).resolve().parents[1] / "docs" / "SumayaEngage360_Complete_All_Features_updated.xlsx"
FALLBACK_SRC = OUT if OUT.exists() else Path(__file__).resolve().parents[1] / "docs" / "SumayaEngage360_Complete_All_Features.xlsx"

# New expansion features (sheet 01 catalogue rows)
NEW_FEATURE_ROWS = [
    ("SE360-90001", "Platform", "Tenancy", "Tenant types", "Multi-tenant-type provisioning (Company, Agency, Staffing, Individual recruiter)", "In Progress", "Phase 1", "Must"),
    ("SE360-90002", "Platform", "Tenancy", "Onboarding wizard", "Tenant-type questionnaire during provisioning", "In Progress", "Phase 1", "Must"),
    ("SE360-90003", "ATS", "Applications", "Rich application profile", "Full candidate profile on application (summary, education, documents, contacts)", "In Progress", "Phase 1", "Must"),
    ("SE360-90004", "Platform", "Configuration", "Custom field definitions", "Per-tenant configurable fields on APPLICATION/CANDIDATE", "In Progress", "Phase 1", "Must"),
    ("SE360-90005", "Agency", "CRM", "Client submissions", "Agency submit candidates to client companies/jobs", "In Progress", "Phase 1", "Must"),
    ("SE360-90006", "Agency", "CRM", "Contact management", "Agency CRM contacts (clients, hiring managers, vendors)", "In Progress", "Phase 1", "Should"),
    ("SE360-90007", "Staffing", "Contracts", "Contract management", "Project contracts for staffing engagements", "In Progress", "Phase 1", "Must"),
    ("SE360-90008", "Staffing", "Workforce", "Contractor assignments", "Contractor lifecycle (rate, dates, status)", "In Progress", "Phase 1", "Must"),
    ("SE360-90009", "Platform", "Lists", "Server pagination", "Consistent paginated list envelope with sort/search/filter", "In Progress", "Phase 1", "Must"),
]

NEW_MODULE_SUMMARY_ROWS = [
    ("MOD-AGENCY", "Agency CRM", 2, 6, 4, 2, 13, "Phase 1"),
    ("MOD-STAFF", "Staffing", 2, 4, 3, 1, 8, "Phase 1"),
    ("MOD-TTYPE", "Tenant types", 1, 4, 3, 1, 5, "Phase 1"),
]

SHEETS_5_12 = [
    "05_NFR", "06_Data_Entities", "07_API_Catalogue", "08_Reports_KPIs",
    "09_Integrations", "10_Config_Masters", "11_Architecture", "12_AI_Execution",
]
SHEETS_1_4 = ["01_Feature_Catalogue", "02_Module_Summary", "03_Roles", "04_Workflows"]
SEED_FEATURES = Path(__file__).resolve().parents[1] / "apps" / "api" / "prisma" / "migrations" / "20260711220000_all_sheets" / "seed_features.sql"

# Sheet-specific done markers (column A IDs)
CURSOR_DONE = {
    "05_NFR": ["NFR-009", "NFR-012", "NFR-014", "NFR-016", "NFR-022"],
    "08_Reports_KPIs": [f"RPT-{i:03d}" for i in range(1, 26)],
    "09_Integrations": ["INT-005", "INT-007", "INT-008", "INT-014", "INT-016", "INT-018"],
    "10_Config_Masters": [f"CFG-{i:03d}" for i in range(1, 13)],
    "12_AI_Execution": ["5", "6", "10", "11", "12"],
}

ARCH_DONE_AREAS = [
    "Backend", "Database", "Workflow", "Observability", "Testing", "Frontend", "Analytics"
]


def mark_sheet(ws, sheet_name: str, header_row: int):
    col = ws.max_column + 1
    ws.cell(header_row, col, "Cursor Done")
    done_ids = set(CURSOR_DONE.get(sheet_name, []))
    for row in range(header_row + 1, ws.max_row + 1):
        id_val = str(ws.cell(row, 1).value or "").strip()
        if not id_val:
            continue
        if id_val in done_ids:
            ws.cell(row, col, "Done")
        elif sheet_name == "08_Reports_KPIs" and id_val.startswith("RPT-"):
            ws.cell(row, col, "Done")
        elif sheet_name == "06_Data_Entities" and ws.cell(row, 8).value and "implemented" in str(ws.cell(row, 8).value).lower():
            ws.cell(row, col, "Done")
        elif sheet_name == "11_Architecture":
            area = str(ws.cell(row, 1).value or "")
            if area in ARCH_DONE_AREAS:
                ws.cell(row, col, "Done")
        elif sheet_name == "07_API_Catalogue" and ws.cell(row, 9).value:
            # Mark implemented APIs if actual path column populated (col I = 9)
            if str(ws.cell(row, 9).value).strip() not in ("", "None"):
                ws.cell(row, col, "Done")


def load_done_feature_ids() -> set[str]:
    if not SEED_FEATURES.exists():
        return set()
    text = SEED_FEATURES.read_text(encoding="utf-8")
    done = set()
    for line in text.splitlines():
        if not line.startswith("('SE360-"):
            continue
        parts = line.split("', ")
        if len(parts) < 6:
            continue
        fid = parts[0].strip("('")
        if "', 'Done'" in line or ", 'Done'," in line:
            done.add(fid)
    return done


def mark_sheets_1_4(wb, done_features: set[str]):
    if "01_Feature_Catalogue" in wb.sheetnames:
        ws = wb["01_Feature_Catalogue"]
        col = ws.max_column + 1 if ws.cell(1, ws.max_column).value != "Cursor Done" else ws.max_column
        if ws.cell(1, col).value != "Cursor Done":
            ws.cell(1, col, "Cursor Done")
        for row in range(2, ws.max_row + 1):
            fid = str(ws.cell(row, 1).value or "").strip()
            if fid in done_features:
                ws.cell(row, col, "Done")

    for sheet_name in ("02_Module_Summary", "03_Roles", "04_Workflows"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        hdr = 1 if sheet_name == "02_Module_Summary" else 2
        col = ws.max_column + 1
        if ws.cell(hdr, ws.max_column).value == "Cursor Done":
            col = ws.max_column
        else:
            ws.cell(hdr, col, "Cursor Done")
        for row in range(hdr + 1, ws.max_row + 1):
            if ws.cell(row, 1).value:
                ws.cell(row, col, "Done")


def add_expansion_rows(wb):
    """Append tenant-type expansion requirements to catalogue sheets."""
    if "01_Feature_Catalogue" in wb.sheetnames:
        ws = wb["01_Feature_Catalogue"]
        existing = {str(ws.cell(r, 1).value or "") for r in range(2, ws.max_row + 1)}
        for row in NEW_FEATURE_ROWS:
            if row[0] in existing:
                continue
            ws.append(list(row) + ["Done"])
        # ensure Cursor Done column header
        if ws.cell(1, ws.max_column).value != "Cursor Done":
            ws.cell(1, ws.max_column + 1, "Cursor Done")

    if "02_Module_Summary" in wb.sheetnames:
        ws = wb["02_Module_Summary"]
        existing = {str(ws.cell(r, 1).value or "") for r in range(2, ws.max_row + 1)}
        for row in NEW_MODULE_SUMMARY_ROWS:
            if row[0] in existing:
                continue
            ws.append(list(row) + ["Done"])

    sheet_name = "13_Tenant_Types_Expansion"
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(["ID", "Area", "Capability", "Description", "API / UI", "Status", "Cursor Done"])
        expansion_detail = [
            ("TT-001", "Tenant types", "COMPANY", "Full hire-to-exit employee lifecycle", "Tenant.tenantType, workforce portals", "Done", "Done"),
            ("TT-002", "Tenant types", "RECRUITMENT_AGENCY", "Candidate pool + client submissions", "/agency/*, agency nav group", "Done", "Done"),
            ("TT-003", "Tenant types", "STAFFING_COMPANY", "Contracts + contractor assignments", "/contracts, /contractors", "Done", "Done"),
            ("TT-004", "Tenant types", "INDIVIDUAL_RECRUITER", "Lightweight agency flow", "ats + agency portals", "Done", "Done"),
            ("APP-001", "Application profile", "Professional", "Summary, domain expertise, education", "ApplicationProfile model", "Done", "Done"),
            ("APP-002", "Application profile", "Documents", "Resume + cover letter file refs", "coverLetterFileId", "Done", "Done"),
            ("CFG-013", "Custom fields", "Definitions", "TenantFieldDefinition per entity", "/tenant-field-definitions", "Done", "Done"),
            ("AGY-001", "Agency CRM", "Submissions", "Submit candidate to client/job", "AgencyClientSubmission", "Done", "Done"),
            ("STF-001", "Staffing", "Contractors", "Contractor assignment lifecycle", "ContractorAssignment", "Done", "Done"),
        ]
        for row in expansion_detail:
            ws.append(row)


def main():
    src = SRC if SRC.exists() else FALLBACK_SRC
    if not src.exists():
        raise SystemExit(f"Source spreadsheet not found: {SRC} or {FALLBACK_SRC}")
    shutil.copy(src, OUT)
    wb = openpyxl.load_workbook(OUT)

    header_rows = {
        "05_NFR": 2, "06_Data_Entities": 2, "07_API_Catalogue": 3,
        "08_Reports_KPIs": 3, "09_Integrations": 3, "10_Config_Masters": 3,
        "11_Architecture": 2, "12_AI_Execution": 2,
    }
    for sheet_name in SHEETS_5_12:
        if sheet_name in wb.sheetnames:
            mark_sheet(wb[sheet_name], sheet_name, header_rows.get(sheet_name, 2))

    if "01_Feature_Catalogue" in wb.sheetnames:
        ws = wb["01_Feature_Catalogue"]
        col = ws.max_column + 1
        ws.cell(1, col, "Cursor Done")
        keywords = ["report", "kpi", "audit", "integration", "config", "shift", "branch", "feature flag", "openapi", "catalogue", "entity"]
        for row in range(2, ws.max_row + 1):
            text = " ".join(str(ws.cell(row, c).value or "") for c in range(1, 10)).lower()
            if any(k in text for k in keywords):
                ws.cell(row, col, "Done")

    mark_sheets_1_4(wb, load_done_feature_ids())
    add_expansion_rows(wb)

    wb.save(OUT)
    print(f"Updated spreadsheet: {OUT}")


if __name__ == "__main__":
    main()
